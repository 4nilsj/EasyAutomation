#!/usr/bin/env python3
"""
Multi-PDF to Excel (XLSX) Converter
===================================
Extract tables and text from multiple PDF files and convert each PDF into a separate, 
professionally formatted worksheet in a single Excel (.xlsx) workbook.

Dependencies:
    pip install pdfplumber pandas openpyxl

Usage:
    # Convert multiple PDFs into one Excel file (each PDF gets its own sheet):
    python pdftoxlsx.py -i file1.pdf file2.pdf file3.pdf -o merged_output.xlsx

    # Convert all PDFs in a directory:
    python pdftoxlsx.py -d /path/to/pdf_folder -o folder_export.xlsx

    # Convert a single PDF:
    python pdftoxlsx.py -i document.pdf
"""

import argparse
import glob
import logging
import os
import re
import sys
from typing import List, Optional, Set, Tuple

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger("PDF2XLSX")


def check_dependencies():
    """Verify required third-party libraries are installed."""
    missing = []
    try:
        import pdfplumber
    except ImportError:
        missing.append("pdfplumber")
    
    try:
        import pandas
    except ImportError:
        missing.append("pandas")
        
    try:
        import openpyxl
    except ImportError:
        missing.append("openpyxl")
        
    if missing:
        logger.error(f"Missing required packages: {', '.join(missing)}")
        logger.info(f"Please install them using: pip install {' '.join(missing)}")
        sys.exit(1)


def parse_page_range(pages_str: str, total_pages: int) -> List[int]:
    """Parse a page range string (e.g., '1-5', '1,3,5', 'all') into 0-based page indices."""
    if pages_str.lower() == 'all':
        return list(range(total_pages))
    
    indices = set()
    parts = pages_str.split(',')
    for part in parts:
        part = part.strip()
        if '-' in part:
            try:
                start, end = part.split('-')
                start_idx = max(1, int(start)) - 1
                end_idx = min(total_pages, int(end))
                for i in range(start_idx, end_idx):
                    indices.add(i)
            except ValueError:
                logger.warning(f"Invalid range format '{part}', ignoring.")
        else:
            try:
                page_num = int(part)
                if 1 <= page_num <= total_pages:
                    indices.add(page_num - 1)
            except ValueError:
                logger.warning(f"Invalid page number '{part}', ignoring.")
                
    return sorted(list(indices))


def clean_table_data(table: List[List[Optional[str]]]) -> List[List[str]]:
    """Clean extracted table cells by stripping whitespace, replacing Nones, and normalizing internal newlines."""
    cleaned = []
    for row in table:
        if not row:
            continue
        cleaned_row = []
        for cell in row:
            if cell is None:
                cleaned_row.append("")
            else:
                # Replace multi-line breaks and excessive whitespace with a single space
                cleaned_cell = re.sub(r'\s+', ' ', str(cell)).strip()
                cleaned_row.append(cleaned_cell)
        if any(c != "" for c in cleaned_row):
            cleaned.append(cleaned_row)
    return cleaned



def sanitize_sheet_name(raw_name: str, used_names: Set[str]) -> str:
    """
    Sanitize string for Excel sheet name:
    - Remove invalid characters: \\ / ? * : [ ]
    - Uses first 10 characters of the PDF file name
    - Enforces uniqueness in used_names set (appends _1, _2 if duplicate 10-char names exist)
    """
    # Remove invalid characters
    clean_name = re.sub(r'[\\/*?:\[\]]', '_', raw_name).strip()
    if not clean_name:
        clean_name = "Sheet"
        
    # Truncate to first 10 characters
    base_name = clean_name[:10]
    candidate = base_name
    
    counter = 1
    # Check case-insensitive uniqueness for Excel compatibility
    lower_used = {n.lower() for n in used_names}
    
    while candidate.lower() in lower_used:
        suffix = f"_{counter}"
        avail_len = 31 - len(suffix)
        candidate = f"{base_name[:avail_len]}{suffix}"
        counter += 1
        
    used_names.add(candidate)
    return candidate



def extract_pdf_data(
    pdf_path: str, 
    page_indices: List[int],
    extract_text_fallback: bool = True
) -> List[dict]:
    """Extract tables and text from specified pages of a PDF."""
    import pdfplumber
    import pandas as pd

    page_data = []

    with pdfplumber.open(pdf_path) as pdf:
        total_pdf_pages = len(pdf.pages)
        
        for idx in page_indices:
            if idx >= total_pdf_pages:
                continue
            page_num = idx + 1
            page = pdf.pages[idx]
            
            raw_tables = page.extract_tables()
            processed_dfs = []
            
            for raw_table in raw_tables:
                cleaned = clean_table_data(raw_table)
                if not cleaned:
                    continue
                
                header = cleaned[0]
                data_rows = cleaned[1:] if len(cleaned) > 1 else []
                
                if data_rows:
                    df = pd.DataFrame(data_rows, columns=header)
                else:
                    df = pd.DataFrame(columns=header)
                    
                processed_dfs.append(df)
            
            extracted_text = ""
            if extract_text_fallback and not processed_dfs:
                extracted_text = page.extract_text() or ""
                
            page_data.append({
                'page': page_num,
                'tables': processed_dfs,
                'text': extracted_text
            })

    return page_data


def format_excel_workbook(xlsx_path: str):
    """
    Apply professional formatting to all worksheets using openpyxl:
    - Auto-adjust column widths
    - Styled headers (Navy background, bold white text)
    - Thin borders and zebra striping for rows
    - Explicit gridlines enablement
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(xlsx_path)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    zebra_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    regular_font = Font(name="Calibri", size=11, color="000000")
    
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    cell_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side
    )

    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True

        if ws.max_row == 0 or ws.max_column == 0:
            continue

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
            is_header = (row_idx == 1)
            
            for cell in row:
                if is_header:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                else:
                    cell.font = regular_font
                    cell.alignment = data_alignment
                    if row_idx % 2 == 0:
                        cell.fill = zebra_fill
                
                cell.border = cell_border

        ws.row_dimensions[1].height = 28
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 20

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            
            adjusted_width = min(max(max_len + 4, 12), 60)
            ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(xlsx_path)
    logger.info(f"Applied styling and auto-formatting across all sheets in '{xlsx_path}'.")


def convert_multiple_pdfs_to_excel(
    pdf_paths: List[str], 
    output_xlsx: str, 
    pages_str: str = "all"
):
    """
    Convert a list of PDF files into a single Excel file, placing each PDF on a new sheet.
    """
    check_dependencies()
    import pdfplumber
    import pandas as pd

    if not pdf_paths:
        logger.error("No PDF files provided for conversion.")
        sys.exit(1)

    used_sheet_names: Set[str] = set()
    writer = pd.ExcelWriter(output_xlsx, engine='openpyxl')
    processed_count = 0

    for pdf_path in pdf_paths:
        if not os.path.exists(pdf_path):
            logger.warning(f"File not found: '{pdf_path}', skipping.")
            continue

        filename_without_ext = os.path.splitext(os.path.basename(pdf_path))[0]
        sheet_name = sanitize_sheet_name(filename_without_ext, used_sheet_names)

        logger.info(f"Processing PDF: '{pdf_path}' -> Sheet: '{sheet_name}'")

        try:
            with pdfplumber.open(pdf_path) as pdf:
                total_pages = len(pdf.pages)
            
            page_indices = parse_page_range(pages_str, total_pages)
            page_data = extract_pdf_data(pdf_path, page_indices)

            # Collect all dataframes or text lines from this PDF
            pdf_dfs = []
            for item in page_data:
                p_num = item['page']
                tables = item['tables']
                text = item['text']

                if tables:
                    for df in tables:
                        df_copy = df.copy()
                        df_copy.insert(0, 'Page', p_num)
                        pdf_dfs.append(df_copy)
                elif text:
                    lines = [line for line in text.strip().split('\n') if line.strip()]
                    df = pd.DataFrame({'Page': p_num, 'Extracted Text': lines})
                    pdf_dfs.append(df)

            if pdf_dfs:
                try:
                    combined_df = pd.concat(pdf_dfs, ignore_index=True)
                    combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
                except Exception:
                    # If columns differ drastically between tables on different pages
                    current_row = 0
                    for df in pdf_dfs:
                        df.to_excel(writer, sheet_name=sheet_name, startrow=current_row, index=False)
                        current_row += len(df) + 3
            else:
                empty_df = pd.DataFrame({"Notice": [f"No tables or text extracted from '{os.path.basename(pdf_path)}'."]})
                empty_df.to_excel(writer, sheet_name=sheet_name, index=False)

            processed_count += 1

        except Exception as e:
            logger.error(f"Error processing '{pdf_path}': {e}")
            empty_df = pd.DataFrame({"Error": [f"Failed to extract content: {e}"]})
            empty_df.to_excel(writer, sheet_name=sheet_name, index=False)

    if processed_count == 0:
        logger.error("No PDFs were successfully processed.")
        sys.exit(1)

    writer.close()
    logger.info(f"Successfully converted {processed_count} PDF file(s) into '{output_xlsx}'.")

    # Format the entire workbook with openpyxl
    format_excel_workbook(output_xlsx)


def collect_pdf_files(input_list: Optional[List[str]], dir_path: Optional[str]) -> List[str]:
    """Gather all valid PDF paths from command line arguments or directory search."""
    pdf_files = []
    
    if input_list:
        for item in input_list:
            # Expand glob patterns if shell did not expand them
            expanded = glob.glob(item)
            if expanded:
                pdf_files.extend(expanded)
            else:
                pdf_files.append(item)
                
    if dir_path:
        if os.path.exists(dir_path):
            dir_pdfs = sorted(glob.glob(os.path.join(dir_path, "*.pdf")))
            pdf_files.extend(dir_pdfs)
        else:
            logger.error(f"Directory not found: '{dir_path}'")
            
    # Filter unique PDF paths while preserving order
    seen = set()
    unique_pdfs = []
    for path in pdf_files:
        abs_path = os.path.abspath(path)
        if abs_path not in seen and path.lower().endswith('.pdf'):
            seen.add(abs_path)
            unique_pdfs.append(path)
            
    return unique_pdfs


def main():
    parser = argparse.ArgumentParser(
        description="Convert multiple PDF files, placing each PDF on a new sheet in a single Excel (.xlsx) workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Convert multiple PDF files into one Excel file (one sheet per PDF):
  python pdftoxlsx.py -i invoice1.pdf invoice2.pdf invoice3.pdf -o invoices_combined.xlsx

  # Convert all PDFs matching a wildcard pattern:
  python pdftoxlsx.py -i *.pdf -o all_pdfs.xlsx

  # Convert all PDFs in a folder:
  python pdftoxlsx.py -d /path/to/pdfs -o folder_report.xlsx
        """
    )
    
    parser.add_argument(
        "-i", "--inputs", 
        nargs="+", 
        help="One or more PDF file paths (e.g. -i doc1.pdf doc2.pdf or *.pdf)."
    )
    parser.add_argument(
        "-d", "--dir", 
        help="Path to a directory containing PDF files."
    )
    parser.add_argument(
        "-o", "--output", 
        default="multi_pdf_export.xlsx", 
        help="Path to output Excel file (default: multi_pdf_export.xlsx)."
    )
    parser.add_argument(
        "-p", "--pages", 
        default="all", 
        help="Page range to process per PDF (e.g., 'all', '1-5', '1,3'). Default: 'all'."
    )

    args = parser.parse_args()

    if not args.inputs and not args.dir:
        parser.error("You must specify either -i/--inputs or -d/--dir.")

    pdf_files = collect_pdf_files(args.inputs, args.dir)
    
    if not pdf_files:
        logger.error("No valid .pdf files were found to convert.")
        sys.exit(1)

    logger.info(f"Found {len(pdf_files)} PDF file(s) to process.")
    
    # If only 1 PDF was passed and default output wasn't changed, name output after the PDF
    output_file = args.output
    if len(pdf_files) == 1 and args.output == "multi_pdf_export.xlsx":
        base_name = os.path.splitext(pdf_files[0])[0]
        output_file = f"{base_name}.xlsx"

    convert_multiple_pdfs_to_excel(
        pdf_paths=pdf_files,
        output_xlsx=output_file,
        pages_str=args.pages
    )


if __name__ == "__main__":
    main()
