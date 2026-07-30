#!/usr/bin/env python3
"""
Merge Excel Sheets into Single Master Sheet
===========================================
Reads a multi-sheet Excel file (.xlsx) and consolidates all data into a single 
master worksheet. Adds a 'Sheet Name' column to indicate the origin of each row.

Dependencies:
    pip install pandas openpyxl

Usage:
    python merge_xlsx_sheets.py -i all_pdfs_combined.xlsx -o master_merged.xlsx
    python merge_xlsx_sheets.py -i input.xlsx --column-name "Source PDF"
"""

import argparse
import logging
import os
import re
import sys
from typing import Optional

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger("MergeExcel")


def check_dependencies():
    """Verify required libraries are installed."""
    missing = []
    try:
        import pandas
    except ImportError:
        missing.append("pandas")
    try:
        import openpyxl
    except ImportError:
        missing.append("openpyxl")

    if missing:
        logger.error(f"Missing required packages: {', '.join(missing)}")
        logger.info(f"Please install them using: pip install {' '.join(missing)}")
        sys.exit(1)


def format_excel_sheet(xlsx_path: str, sheet_name: str = "Merged_Data"):
    """
    Apply professional formatting using openpyxl:
    - Styled headers (Navy blue background, bold white text)
    - Zebra striping (alternating row shading)
    - Auto-adjusted column widths
    - Thin borders and gridlines
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(xlsx_path)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    zebra_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    regular_font = Font(name="Calibri", size=11, color="000000")
    
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    cell_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side
    )

    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True

        if ws.max_row > 0 and ws.max_column > 0:
            # Format rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
                is_header = (row_idx == 1)
                
                for cell in row:
                    if is_header:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                    else:
                        cell.font = regular_font
                        cell.alignment = data_alignment
                        if row_idx % 2 == 0:
                            cell.fill = zebra_fill
                    
                    cell.border = cell_border

            ws.row_dimensions[1].height = 28
            for r in range(2, ws.max_row + 1):
                ws.row_dimensions[r].height = 20

            # Auto-adjust column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or "")
                    lines = val_str.split('\n')
                    for line in lines:
                        if len(line) > max_len:
                            max_len = len(line)
                
                adjusted_width = min(max(max_len + 4, 12), 60)
                ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(xlsx_path)
    logger.info(f"Applied styling and auto-formatting to '{xlsx_path}'")


def merge_excel_sheets(
    input_xlsx: str, 
    output_xlsx: Optional[str] = None,
    sheet_column_name: str = "Sheet Name"
) -> str:
    """
    Reads all sheets from input_xlsx, adds a sheet identification column, 
    concatenates all data into a single sheet, and writes to output_xlsx.
    """
    check_dependencies()
    import pandas as pd

    if not os.path.exists(input_xlsx):
        raise FileNotFoundError(f"Input file not found: '{input_xlsx}'")

    if not output_xlsx:
        base, ext = os.path.splitext(input_xlsx)
        output_xlsx = f"{base}_single_sheet{ext}"

    logger.info(f"Reading multi-sheet Excel file: '{input_xlsx}'")
    # Read all sheets into a dict of {sheet_name: DataFrame}
    sheets_dict = pd.read_excel(input_xlsx, sheet_name=None)
    
    if not sheets_dict:
        logger.error("No sheets found in the provided Excel file.")
        sys.exit(1)

    logger.info(f"Found {len(sheets_dict)} sheet(s) in input file.")

    combined_dfs = []
    total_rows = 0

    for sheet_name, df in sheets_dict.items():
        if df.empty:
            logger.info(f"Skipping empty sheet: '{sheet_name}'")
            continue

        df_copy = df.copy()
        # Normalize column headers (convert internal newlines/spaces to single space)
        df_copy.columns = [re.sub(r'\s+', ' ', str(c)).strip() for c in df_copy.columns]
        # Insert Sheet Name column as the first column
        df_copy.insert(0, sheet_column_name, sheet_name)
        combined_dfs.append(df_copy)
        total_rows += len(df_copy)

    if not combined_dfs:
        logger.error("All sheets in the file were empty. Nothing to merge.")
        sys.exit(1)

    logger.info(f"Merging {len(combined_dfs)} sheet(s) with total {total_rows} row(s)...")
    
    # Combine into single DataFrame
    master_df = pd.concat(combined_dfs, ignore_index=True)

    master_sheet_title = "Merged_All_Sheets"
    
    # Write to Excel
    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        master_df.to_excel(writer, sheet_name=master_sheet_title, index=False)

    logger.info(f"Successfully saved merged sheet to '{output_xlsx}'")
    
    # Apply openpyxl styling
    format_excel_sheet(output_xlsx, sheet_name=master_sheet_title)

    return output_xlsx


def main():
    parser = argparse.ArgumentParser(
        description="Merge all sheets of an Excel (.xlsx) workbook into a single sheet with a Sheet Name column."
    )
    parser.add_argument(
        "-i", "--input", 
        required=True, 
        help="Path to input multi-sheet Excel file."
    )
    parser.add_argument(
        "-o", "--output", 
        default=None, 
        help="Path to output Excel file (default: <input>_single_sheet.xlsx)."
    )
    parser.add_argument(
        "-c", "--column-name", 
        default="Sheet Name", 
        help="Name of the column storing sheet names (default: 'Sheet Name')."
    )

    args = parser.parse_args()

    try:
        merge_excel_sheets(
            input_xlsx=args.input,
            output_xlsx=args.output,
            sheet_column_name=args.column_name
        )
    except Exception as e:
        logger.error(f"Error during merging: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
